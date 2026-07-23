#!/usr/bin/env python3
"""
社服部陌拜跟踪系统 — 本地服务器
启动后访问 http://localhost:8080
数据持久化到 data.json + Excel导入导出
"""

import http.server
import json
import os
import io
import urllib.parse
import mimetypes
import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

PORT = int(os.environ.get('PORT', 8080))
DATA_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_FILE = os.path.join(DATA_DIR, 'data.json')
HTML_FILE = os.path.join(DATA_DIR, '市场陌拜跟踪系统.html')

# Colors
DARK_BLUE = '101843'
ACCENT = 'FF6B35'
LIGHT_BG = 'F5F6FA'
HEADER_FONT = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
HEADER_FILL = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type='solid')
ACCENT_FILL = PatternFill(start_color=ACCENT, end_color=ACCENT, fill_type='solid')
LIGHT_FILL = PatternFill(start_color=LIGHT_BG, end_color=LIGHT_BG, fill_type='solid')
THIN_BORDER = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC'),
)

def load_data():
    if os.path.exists(DATA_FILE):
        try:
            with open(DATA_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except (json.JSONDecodeError, IOError):
            return None
    return None

def save_data(data):
    with open(DATA_FILE, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

def get_person_name(data, pid):
    for p in data.get('personnel', []):
        if p['id'] == pid:
            return p['name']
    return pid

def style_header(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = THIN_BORDER

def style_data(ws, row, cols, alt=False):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = Font(name='微软雅黑', size=10)
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center', vertical='center')
        if alt:
            cell.fill = LIGHT_FILL

# ===== EXPORT FUNCTIONS =====

def export_performance(data, month=None):
    """导出月度业绩汇总"""
    wb = Workbook()
    ws = wb.active
    ws.title = '业绩汇总'

    # Title
    month_str = f'{month}月' if month else '全年'
    ws.merge_cells('A1:G1')
    ws['A1'] = f'京顺医院社服部 {month_str}业绩汇总表'
    ws['A1'].font = Font(name='微软雅黑', bold=True, size=14, color=DARK_BLUE)
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[1].height = 30

    # Headers
    headers = ['姓名', '分组', '负责网点', month_str + '业绩(元)', '拜访次数', '目标数', '备注']
    for i, h in enumerate(headers, 1):
        ws.cell(row=3, column=i, value=h)
    style_header(ws, 3, len(headers))

    # Data
    records = data.get('performance', {}).get('records', [])
    visits = data.get('visits', [])
    targets = data.get('targets', [])
    personnel = data.get('personnel', [])

    person_data = {}
    for p in personnel:
        if p.get('group') == '—' or not p.get('group'):
            continue
        pid = p['id']
        person_data[pid] = {
            'name': p['name'],
            'group': p.get('group', ''),
            'sites': p.get('sites', ''),
            'perf': 0,
            'visits': 0,
            'targets': 0,
        }

    for r in records:
        pid = r.get('personId')
        if pid in person_data:
            if month is None or r.get('month') == month:
                person_data[pid]['perf'] += r.get('amount', 0)

    for v in visits:
        pid = v.get('personId')
        if pid in person_data:
            if month is None:
                person_data[pid]['visits'] += 1
            elif v.get('date', '').startswith(f'2026-{month:02d}'):
                person_data[pid]['visits'] += 1

    for t in targets:
        pid = t.get('assignee')
        if pid in person_data:
            person_data[pid]['targets'] += 1

    row = 4
    total_perf = 0
    total_visits = 0
    for pid, pd in sorted(person_data.items(), key=lambda x: -x[1]['perf']):
        vals = [pd['name'], pd['group'], pd['sites'],
                round(pd['perf'], 2), pd['visits'], pd['targets'], '']
        for i, v in enumerate(vals, 1):
            ws.cell(row=row, column=i, value=v)
        style_data(ws, row, len(headers), alt=(row % 2 == 0))
        total_perf += pd['perf']
        total_visits += pd['visits']
        row += 1

    # Total row
    ws.cell(row=row, column=1, value='合计')
    ws.cell(row=row, column=4, value=round(total_perf, 2))
    ws.cell(row=row, column=5, value=total_visits)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = Font(name='微软雅黑', bold=True, size=10, color=DARK_BLUE)
        cell.fill = ACCENT_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center')

    # Column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 20

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

def export_visits(data):
    """导出拜访记录"""
    wb = Workbook()
    ws = wb.active
    ws.title = '拜访记录'

    ws.merge_cells('A1:F1')
    ws['A1'] = '京顺医院社服部 拜访记录导出'
    ws['A1'].font = Font(name='微软雅黑', bold=True, size=14, color=DARK_BLUE)
    ws['A1'].alignment = Alignment(horizontal='center')

    headers = ['日期', '陌拜人员', '目标', '阶段', '关键信息', '下次跟进']
    for i, h in enumerate(headers, 1):
        ws.cell(row=3, column=i, value=h)
    style_header(ws, 3, len(headers))

    STAGE_NAMES = {1: '信息收集', 2: '建立链接', 3: '关系递进', 4: '合作转诊'}
    visits = data.get('visits', [])
    visits.sort(key=lambda v: v.get('date', ''), reverse=True)

    row = 4
    for v in visits:
        pid = v.get('personId', '')
        tid = v.get('targetId', '')
        pname = get_person_name(data, pid)
        tname = ''
        for t in data.get('targets', []):
            if t['id'] == tid:
                tname = t['name']
                break
        stage = STAGE_NAMES.get(v.get('stage', 1), '')

        ws.cell(row=row, column=1, value=v.get('date', ''))
        ws.cell(row=row, column=2, value=pname)
        ws.cell(row=row, column=3, value=tname)
        ws.cell(row=row, column=4, value=stage)
        ws.cell(row=row, column=5, value=v.get('note', ''))
        ws.cell(row=row, column=6, value=v.get('nextDate', ''))
        style_data(ws, row, len(headers), alt=(row % 2 == 0))
        row += 1

    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 24
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 40
    ws.column_dimensions['F'].width = 14

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

def export_targets(data):
    """导出陌拜目标清单"""
    wb = Workbook()
    ws = wb.active
    ws.title = '陌拜目标'

    ws.merge_cells('A1:G1')
    ws['A1'] = '京顺医院社服部 陌拜目标清单'
    ws['A1'].font = Font(name='微软雅黑', bold=True, size=14, color=DARK_BLUE)
    ws['A1'].alignment = Alignment(horizontal='center')

    headers = ['目标名称', '类型', '负责人', '阶段', '科室', '备注/基础信息']
    for i, h in enumerate(headers, 1):
        ws.cell(row=3, column=i, value=h)
    style_header(ws, 3, len(headers))

    targets = data.get('targets', [])
    STAGE_NAMES = {1: '①信息收集', 2: '②建立链接', 3: '③关系递进', 4: '④合作转诊'}

    row = 4
    for t in targets:
        pid = t.get('assignee', '')
        pname = get_person_name(data, pid)
        stage = STAGE_NAMES.get(t.get('stage', 1), '')

        ws.cell(row=row, column=1, value=t.get('name', ''))
        ws.cell(row=row, column=2, value=t.get('type', ''))
        ws.cell(row=row, column=3, value=pname)
        ws.cell(row=row, column=4, value=stage)
        ws.cell(row=row, column=5, value=t.get('dept', ''))
        ws.cell(row=row, column=6, value=t.get('note', ''))
        style_data(ws, row, len(headers), alt=(row % 2 == 0))
        row += 1

    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 16
    ws.column_dimensions['F'].width = 50

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

# ===== IMPORT FUNCTION =====

def import_performance(data, excel_bytes, filename):
    """从Excel导入业绩数据"""
    imported = 0
    errors = []

    try:
        wb = load_workbook(io.BytesIO(excel_bytes), data_only=True)
        ws = wb.active

        # Build name->id map
        name_map = {}
        for p in data.get('personnel', []):
            name_map[p['name']] = p['id']

        for row in ws.iter_rows(min_row=4, values_only=True):
            name = str(row[0]).strip() if row[0] else ''
            amount = row[3] if len(row) > 3 and row[3] else 0
            month = row[4] if len(row) > 4 and row[4] else None

            if not name or not amount:
                continue

            pid = name_map.get(name)
            if not pid:
                errors.append(f'未找到人员: {name}')
                continue

            try:
                amount = float(amount)
            except (ValueError, TypeError):
                errors.append(f'金额格式错误: {name} - {amount}')
                continue

            if month is None:
                month = datetime.datetime.now().month

            try:
                month = int(month)
            except (ValueError, TypeError):
                month = datetime.datetime.now().month

            data['performance']['records'].append({
                'id': f'imp_{datetime.datetime.now().timestamp()}_{imported}',
                'personId': pid,
                'month': month,
                'amount': amount,
                'source': 'Excel导入',
                'date': datetime.date.today().isoformat(),
                'note': f'从{filename}导入',
            })
            imported += 1

    except Exception as e:
        errors.append(f'文件解析失败: {str(e)}')

    return imported, errors


class Handler(http.server.SimpleHTTPRequestHandler):
    def do_GET(self):
        parsed = urllib.parse.urlparse(self.path)
        query = urllib.parse.parse_qs(parsed.query)
        path = parsed.path

        data = load_data() or {}

        # API: 获取数据
        if path == '/api/data':
            self.send_json(200, data if data else {})
            return

        # API: 导出业绩
        if path == '/api/export/performance':
            month = int(query.get('month', [0])[0]) or None
            buf = export_performance(data, month)
            filename = f'业绩汇总表_{month or "全年"}月.xlsx'
            self.send_file(buf, filename)
            return

        # API: 导出拜访
        if path == '/api/export/visits':
            buf = export_visits(data)
            self.send_file(buf, '拜访记录导出.xlsx')
            return

        # API: 导出目标
        if path == '/api/export/targets':
            buf = export_targets(data)
            self.send_file(buf, '陌拜目标清单.xlsx')
            return

        # API: 导出完整月报
        if path == '/api/export/monthly-report':
            month = int(query.get('month', [datetime.datetime.now().month])[0])
            wb = Workbook()
            # Performance sheet
            ws1 = wb.active
            ws1.title = '业绩汇总'
            perf_buf = export_performance(data, month)
            perf_wb = load_workbook(perf_buf)
            perf_ws = perf_wb.active
            for row in perf_ws.iter_rows(min_row=1, max_row=perf_ws.max_row, values_only=True):
                ws1.append(row)

            # Visits sheet
            ws2 = wb.create_sheet('拜访记录')
            vis_buf = export_visits(data)
            vis_wb = load_workbook(vis_buf)
            vis_ws = vis_wb.active
            for row in vis_ws.iter_rows(min_row=1, max_row=vis_ws.max_row, values_only=True):
                ws2.append(row)

            # Targets sheet
            ws3 = wb.create_sheet('陌拜目标')
            tgt_buf = export_targets(data)
            tgt_wb = load_workbook(tgt_buf)
            tgt_ws = tgt_wb.active
            for row in tgt_ws.iter_rows(min_row=1, max_row=tgt_ws.max_row, values_only=True):
                ws3.append(row)

            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            self.send_file(buf, f'社服部{month}月完整报表.xlsx')
            return

        super().do_GET()

    def do_POST(self):
        parsed = urllib.parse.urlparse(self.path)
        path = parsed.path

        data = load_data() or {}

        # API: 保存数据
        if path == '/api/data':
            content_len = int(self.headers.get('Content-Length', 0))
            body = self.rfile.read(content_len)
            try:
                new_data = json.loads(body.decode('utf-8'))
                save_data(new_data)
                self.send_json(200, {'status': 'ok', 'message': '数据已保存'})
            except Exception as e:
                self.send_json(400, {'status': 'error', 'message': str(e)})
            return

        # API: 导入业绩
        if path == '/api/import/performance':
            content_type = self.headers.get('Content-Type', '')
            if 'multipart/form-data' not in content_type:
                self.send_json(400, {'status': 'error', 'message': '请上传Excel文件'})
                return

            # Parse multipart
            boundary = content_type.split('boundary=')[1].strip()
            body = self.rfile.read(int(self.headers.get('Content-Length', 0)))

            # Extract file data
            parts = body.split(b'--' + boundary.encode())
            excel_data = None
            filename = 'import.xlsx'

            for part in parts:
                if b'Content-Disposition' in part and b'filename=' in part:
                    # Extract filename
                    fn_start = part.find(b'filename="') + 10
                    fn_end = part.find(b'"', fn_start)
                    if fn_start > 9:
                        filename = part[fn_start:fn_end].decode('utf-8', errors='replace')

                    # Find empty line before file content
                    hdr_end = part.find(b'\r\n\r\n')
                    if hdr_end > 0:
                        excel_data = part[hdr_end + 4:]
                        # Remove trailing boundary markers
                        excel_data = excel_data.rstrip(b'\r\n--')

            if excel_data:
                imported, errors = import_performance(data, excel_data, filename)
                save_data(data)
                self.send_json(200, {
                    'status': 'ok',
                    'imported': imported,
                    'errors': errors,
                    'message': f'成功导入 {imported} 条记录' + (f'，{len(errors)} 条错误' if errors else ''),
                })
            else:
                self.send_json(400, {'status': 'error', 'message': '未找到Excel文件内容'})
            return

        self.send_json(404, {'status': 'error', 'message': '接口不存在'})

    def do_OPTIONS(self):
        self.send_response(200)
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Access-Control-Allow-Methods', 'GET, POST, OPTIONS')
        self.send_header('Access-Control-Allow-Headers', 'Content-Type')
        self.end_headers()

    def send_json(self, code, data):
        self.send_response(code)
        self.send_header('Content-Type', 'application/json; charset=utf-8')
        self.send_header('Access-Control-Allow-Origin', '*')
        self.end_headers()
        self.wfile.write(json.dumps(data, ensure_ascii=False).encode('utf-8'))

    def send_file(self, buf, filename):
        self.send_response(200)
        self.send_header('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        self.send_header('Content-Disposition', f'attachment; filename="{urllib.parse.quote(filename)}"')
        self.send_header('Access-Control-Allow-Origin', '*')
        self.end_headers()
        self.wfile.write(buf.read())

    def translate_path(self, path):
        path = super().translate_path(path)
        return path


if __name__ == '__main__':
    if not os.path.exists(HTML_FILE):
        print(f'❌ 未找到HTML文件：{HTML_FILE}')
        exit(1)

    os.chdir(DATA_DIR)
    print(f'🚀 社服部陌拜跟踪系统 服务器启动')
    print(f'   📁 数据文件：{DATA_FILE}')
    print(f'   📊 导出：业绩/拜访/目标 Excel')
    print(f'   📥 导入：Excel 业绩数据')
    print(f'   🌐 访问地址：http://localhost:{PORT}')
    print(f'   ⌨️  Ctrl+C 停止服务器\n')

    server = http.server.HTTPServer(('0.0.0.0', PORT), Handler)
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print('\n🛑 服务器已停止')
        server.server_close()
